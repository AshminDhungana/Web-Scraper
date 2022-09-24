"""Microsoft Excel Read/Write Operation using openpyl"""

from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl import Workbook
import os.path
import re

class ExcelIO:
    def __init__(self,filename = '', operation = '') -> None:
        self.filename = filename
        self.operation = operation

    #Returns True if file extension matches else returns false
    @classmethod
    def FileCheck(self, filename):
        result = re.match("^.*\.(xlsx|xlsm|xltx|xltm)+$", filename)   
        return result != None

    #Checks if file name is a path or filename and returns filename only
    @classmethod
    def FileNameConversion(self,filename):
        result = re.fullmatch(r".*\\.*", filename) 
        if result != None:
            return os.path.basename(filename)
        else:
            return filename


    #Loads and Selects Sheet from excel file
    @classmethod
    def ExcelLoad(self, filename):
        running3 = True
        while running3:
            try:
                wb = load_workbook(filename)
                raise UnboundLocalError 
            except FileNotFoundError:
                filename = input("File not found, Enter file name or path correctly: ").strip()
            except InvalidFileException:
                print("File format not supported, support for only .xlsx,.xlsm,.xltx,.xltm ")
                filename = input("Enter file name or path correctly: ").strip()
            except UnboundLocalError:
                running3 = False  
            else: 
                pass                  
        print(f'Sheets are: {wb.sheetnames}')
        running = True
        while running:
            try:
                SelectSheet = int(input("Enter the sheet Number, 1 for Sheet 1: ").strip())
                raise UnboundLocalError
            except ValueError:
                print("Invalid input, Only Numeric value allowed, Try again.")
            except UnboundLocalError:
                if SelectSheet > len(wb.sheetnames):
                    print(f"Invalid input, input should be between 0 and {len(wb.sheetnames)} ")
                    running = True
                else:
                    running = False
        ws = wb[wb.sheetnames[SelectSheet - 1]]
        return ws
    
    #Creates a new workbook and saves as filename
    @classmethod
    def ExcelCreate(self, filename='defult.xlsx', operation = ''):
        wb = Workbook()
        ws1 = wb.active
        running = True
        while running:
            operation = input("Change File Operation: 0 to Save, 1 to create sheet and 2 to Exit. ").strip()
            match operation:
                case '':
                    pass
                case '0':      
                    wb.save(filename)
                case '1':
                    wb.create_sheet()
                case '2':
                    running = False
        return True  


          

        



